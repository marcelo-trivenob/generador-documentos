import streamlit as st
import os
import re
import shutil
import zipfile
from datetime import datetime, date
from lxml import etree
import io

# ─────────────────────────────────────────────
#  NAMESPACES
# ─────────────────────────────────────────────
XDR_NS = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
SS_NS  = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

# ─────────────────────────────────────────────
#  HELPERS GENERALES
# ─────────────────────────────────────────────

def formatear_valor(valor):
    if valor is None:
        return ""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    s = str(valor).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s


def encontrar_firma(nombre_valor, firmas_dict, log_fn=None):
    if not nombre_valor or not firmas_dict:
        return None, None
    nombre_buscar = str(nombre_valor).strip().lower()
    for nombre_archivo, img_bytes in firmas_dict.items():
        nombre_sin_ext = os.path.splitext(nombre_archivo)[0].strip().lower()
        if nombre_sin_ext == nombre_buscar:
            ext = os.path.splitext(nombre_archivo)[1].lower()
            return img_bytes, ext
    if log_fn and nombre_valor:
        disponibles = list(firmas_dict.keys())[:5]
        log_fn(f"   ⚠ Sin firma: '{nombre_valor}' | Ejemplos: {disponibles}")
    return None, None


# ─────────────────────────────────────────────
#  XLSX
# ─────────────────────────────────────────────

def parse_cell_ref(ref):
    m = re.match(r'^([A-Za-z]+)(\d+)$', ref)
    if not m:
        raise ValueError(f"Referencia de celda inválida: {ref}")
    return m.group(1).upper(), int(m.group(2))


def set_cell_in_xml(sheet_xml: bytes, cell_ref: str, value) -> bytes:
    root = etree.fromstring(sheet_xml)
    col_letter, row_num = parse_cell_ref(cell_ref)
    sheet_data = root.find(f'{{{SS_NS}}}sheetData')
    if sheet_data is None:
        return sheet_xml
    target_row = None
    for row_elem in sheet_data:
        if row_elem.get('r') and int(row_elem.get('r')) == row_num:
            target_row = row_elem
            break
    if target_row is None:
        target_row = etree.SubElement(sheet_data, f'{{{SS_NS}}}row')
        target_row.set('r', str(row_num))
    target_cell = None
    for c_elem in target_row:
        if c_elem.get('r') == cell_ref:
            target_cell = c_elem
            break
    if target_cell is None:
        target_cell = etree.SubElement(target_row, f'{{{SS_NS}}}c')
        target_cell.set('r', cell_ref)
    for child in list(target_cell):
        target_cell.remove(child)
    str_val = formatear_valor(value)
    target_cell.set('t', 'inlineStr')
    is_elem = etree.SubElement(target_cell, f'{{{SS_NS}}}is')
    t_elem = etree.SubElement(is_elem, f'{{{SS_NS}}}t')
    t_elem.text = str_val
    return etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)


def buscar_marcadores_en_xlsx(plantilla_bytes):
    resultado = {}
    firma_info = {}
    with zipfile.ZipFile(io.BytesIO(plantilla_bytes), 'r') as z:
        sheet_names = [n for n in z.namelist() if re.match(r'xl/worksheets/sheet\d+\.xml$', n)]
        for sn in sheet_names:
            xml_bytes = z.read(sn)
            root = etree.fromstring(xml_bytes)
            marcadores_hoja = {}
            for c_elem in root.iter(f'{{{SS_NS}}}c'):
                ref = c_elem.get('r')
                if not ref:
                    continue
                for t_el in c_elem.iter(f'{{{SS_NS}}}t'):
                    if t_el.text and '{{' in t_el.text:
                        campos = re.findall(r'\{\{(\w+)\}\}', t_el.text)
                        for campo in campos:
                            if campo == 'firma':
                                firma_info[sn] = ref
                            else:
                                marcadores_hoja[ref] = campo
            resultado[sn] = marcadores_hoja
    return resultado, firma_info


def _zip_copy_preservando_formato(zin, archivos_modificados):
    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w') as zout:
        for info in zin.infolist():
            if info.filename in archivos_modificados:
                zout.writestr(
                    zipfile.ZipInfo(info.filename),
                    archivos_modificados[info.filename],
                    compress_type=zipfile.ZIP_DEFLATED
                )
            else:
                data = zin.read(info.filename)
                zi = zipfile.ZipInfo(info.filename)
                zi.compress_type = info.compress_type
                zout.writestr(zi, data)
    return out.getvalue()


def llenar_xlsx_zip(plantilla_bytes, datos):
    SS_STR = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    zin_buf = io.BytesIO(plantilla_bytes)
    with zipfile.ZipFile(zin_buf, 'r') as zin:
        archivos = {n: zin.read(n) for n in zin.namelist()}
    modificados = {}
    if 'xl/sharedStrings.xml' in archivos:
        root = etree.fromstring(archivos['xl/sharedStrings.xml'])
        cambio = False
        for si in root.findall(f'{{{SS_STR}}}si'):
            nodos_t = list(si.iter(f'{{{SS_STR}}}t'))
            texto = ''.join(t.text or '' for t in nodos_t)
            if '{{' not in texto:
                continue
            nuevo_texto = texto
            for campo, valor in datos.items():
                if campo == 'firma':
                    continue
                marcador = '{{' + campo + '}}'
                if marcador in nuevo_texto:
                    nuevo_texto = nuevo_texto.replace(marcador, formatear_valor(valor))
            if nuevo_texto != texto:
                cambio = True
                if nodos_t:
                    nodos_t[0].text = nuevo_texto
                    for t in nodos_t[1:]:
                        t.text = ''
        if cambio:
            modificados['xl/sharedStrings.xml'] = etree.tostring(
                root, xml_declaration=True, encoding='UTF-8', standalone=True)
    marcadores, _ = buscar_marcadores_en_xlsx(plantilla_bytes)
    for sheet_name, celdas in marcadores.items():
        if not celdas:
            continue
        sheet_xml = archivos[sheet_name]
        for cell_ref, campo in celdas.items():
            valor = datos.get(campo, "")
            sheet_xml = set_cell_in_xml(sheet_xml, cell_ref, valor)
        modificados[sheet_name] = sheet_xml
    zin_buf.seek(0)
    with zipfile.ZipFile(zin_buf, 'r') as zin:
        return _zip_copy_preservando_formato(zin, modificados)


def _extraer_posicion_anchor(anchor_elem):
    from_el = anchor_elem.find(f'{{{XDR_NS}}}from')
    to_el   = anchor_elem.find(f'{{{XDR_NS}}}to')
    if from_el is None:
        return None
    def safe_int(el, tag):
        v = el.findtext(f'{{{XDR_NS}}}{tag}')
        return int(v) if v is not None else 0
    pos = {
        'col':    safe_int(from_el, 'col'),
        'colOff': safe_int(from_el, 'colOff'),
        'row':    safe_int(from_el, 'row'),
        'rowOff': safe_int(from_el, 'rowOff'),
    }
    if to_el is not None:
        pos['to_col']    = safe_int(to_el, 'col')
        pos['to_colOff'] = safe_int(to_el, 'colOff')
        pos['to_row']    = safe_int(to_el, 'row')
        pos['to_rowOff'] = safe_int(to_el, 'rowOff')
    return pos


def insertar_firmas_en_drawing_xlsx(xlsx_bytes, img_bytes, img_ext, datos, log_fn=None):
    NS_A     = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    NS_R     = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    REL_NS   = 'http://schemas.openxmlformats.org/package/2006/relationships'
    REL_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/image'
    CT_NS    = 'http://schemas.openxmlformats.org/package/2006/content-types'
    NS_PIC   = 'http://schemas.openxmlformats.org/drawingml/2006/picture'
    if img_ext and not img_ext.startswith('.'):
        img_ext = '.' + img_ext
    zin_buf = io.BytesIO(xlsx_bytes)
    with zipfile.ZipFile(zin_buf, 'r') as zin:
        archivos = {n: zin.read(n) for n in zin.namelist()}
    drawing_names = [n for n in archivos if re.match(r'xl/drawings/drawing\d+\.xml$', n)]
    if not drawing_names:
        zin_buf.seek(0)
        with zipfile.ZipFile(zin_buf, 'r') as zin:
            return _zip_copy_preservando_formato(zin, {})
    mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg', '.bmp': 'image/bmp', '.tiff': 'image/tiff'}
    img_mime = mime_map.get(img_ext, 'image/png') if img_ext else 'image/png'
    img_internal_name = f"firma_img{img_ext}" if img_ext else None
    img_zip_path = f"xl/media/{img_internal_name}" if img_internal_name else None
    modificados = {}
    for drawing_name in drawing_names:
        root = etree.fromstring(archivos[drawing_name])
        for t_el in root.iter(f'{{{NS_A}}}t'):
            if not t_el.text or '{{' not in t_el.text:
                continue
            for campo, valor in datos.items():
                if campo == 'firma':
                    continue
                marcador = '{{' + campo + '}}'
                if marcador in t_el.text:
                    t_el.text = t_el.text.replace(marcador, formatear_valor(valor))
        anchors_firma = []
        for anchor in root.findall(f'{{{XDR_NS}}}twoCellAnchor') + root.findall(f'{{{XDR_NS}}}oneCellAnchor'):
            for t_el in anchor.iter(f'{{{NS_A}}}t'):
                if t_el.text and '{{firma}}' in t_el.text:
                    anchors_firma.append(anchor)
                    break
        if anchors_firma and img_bytes and img_ext:
            drawing_basename = os.path.basename(drawing_name)
            rels_name = f"xl/drawings/_rels/{drawing_basename}.rels"
            img_target = f"../media/{img_internal_name}"
            modificados[img_zip_path] = img_bytes
            ct_root = etree.fromstring(archivos['[Content_Types].xml'])
            extension = img_ext.lstrip('.')
            existentes = {e.get('Extension') for e in ct_root.findall(f'{{{CT_NS}}}Default')}
            if extension not in existentes:
                nd = etree.SubElement(ct_root, f'{{{CT_NS}}}Default')
                nd.set('Extension', extension)
                nd.set('ContentType', img_mime)
            modificados['[Content_Types].xml'] = etree.tostring(
                ct_root, xml_declaration=True, encoding='UTF-8', standalone=True)
            if rels_name in archivos:
                rels_root = etree.fromstring(archivos[rels_name])
            else:
                rels_root = etree.fromstring(
                    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>')
            rel_id = 'rFirma1'
            ids_existentes = {e.get('Id') for e in rels_root.findall(f'{{{REL_NS}}}Relationship')}
            if rel_id not in ids_existentes:
                new_rel = etree.SubElement(rels_root, f'{{{REL_NS}}}Relationship')
                new_rel.set('Id', rel_id)
                new_rel.set('Type', REL_TYPE)
                new_rel.set('Target', img_target)
            modificados[rels_name] = etree.tostring(
                rels_root, xml_declaration=True, encoding='UTF-8', standalone=True)
            for anchor in anchors_firma:
                pos = _extraer_posicion_anchor(anchor)
                if pos is None:
                    continue
                to_col    = pos.get('to_col',    pos['col'] + 2)
                to_colOff = pos.get('to_colOff', 0)
                to_row    = pos.get('to_row',    pos['row'] + 3)
                to_rowOff = pos.get('to_rowOff', 0)
                from_el_orig = anchor.find(f'{{{XDR_NS}}}from')
                for child in list(anchor):
                    anchor.remove(child)
                if from_el_orig is not None:
                    anchor.append(from_el_orig)
                to_el = etree.SubElement(anchor, f'{{{XDR_NS}}}to')
                for tag, val in [('col', to_col), ('colOff', to_colOff),
                                  ('row', to_row), ('rowOff', to_rowOff)]:
                    child = etree.SubElement(to_el, f'{{{XDR_NS}}}{tag}')
                    child.text = str(val)
                pic_id = 100
                pic_xml = (
                    f'<xdr:pic xmlns:xdr="{XDR_NS}" xmlns:a="{NS_A}" xmlns:r="{NS_R}" xmlns:pic="{NS_PIC}">'
                    f'<xdr:nvPicPr>'
                    f'<xdr:cNvPr id="{pic_id}" name="firma_img"/>'
                    f'<xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr>'
                    f'</xdr:nvPicPr>'
                    f'<xdr:blipFill>'
                    f'<a:blip r:embed="{rel_id}"/>'
                    f'<a:stretch><a:fillRect/></a:stretch>'
                    f'</xdr:blipFill>'
                    f'<xdr:spPr>'
                    f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="1" cy="1"/></a:xfrm>'
                    f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                    f'</xdr:spPr>'
                    f'</xdr:pic>'
                )
                anchor.append(etree.fromstring(pic_xml))
                etree.SubElement(anchor, f'{{{XDR_NS}}}clientData')
        modificados[drawing_name] = etree.tostring(
            root, xml_declaration=True, encoding='UTF-8', standalone=True)
    zin_buf.seek(0)
    with zipfile.ZipFile(zin_buf, 'r') as zin:
        return _zip_copy_preservando_formato(zin, modificados)


# ─────────────────────────────────────────────
#  DOCX
# ─────────────────────────────────────────────

def llenar_docx(plantilla_bytes, datos, img_bytes, img_ext):
    NS_W   = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
    NS_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    NS_WP  = 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing'
    NS_R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    REL_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/image'
    CT_NS  = 'http://schemas.openxmlformats.org/package/2006/content-types'
    NS_PIC = 'http://schemas.openxmlformats.org/drawingml/2006/picture'
    if img_ext and not img_ext.startswith('.'):
        img_ext = '.' + img_ext
    zin_buf = io.BytesIO(plantilla_bytes)
    with zipfile.ZipFile(zin_buf, 'r') as zin:
        archivos = {n: zin.read(n) for n in zin.namelist()}
    modificados = {}
    root = etree.fromstring(archivos['word/document.xml'])
    for para in root.iter(f'{{{NS_W}}}p'):
        runs = para.findall(f'.//{{{NS_W}}}r')
        if not runs:
            continue
        texto_completo = ''.join(
            (r.findtext(f'{{{NS_W}}}t') or '') for r in runs
        )
        if '{{' not in texto_completo:
            continue
        nuevo_texto = texto_completo
        for campo, valor in datos.items():
            if campo == 'firma':
                continue
            marcador = '{{' + campo + '}}'
            if marcador in nuevo_texto:
                nuevo_texto = nuevo_texto.replace(marcador, formatear_valor(valor))
        if nuevo_texto == texto_completo:
            continue
        primer_run = runs[0]
        t_elem = primer_run.find(f'{{{NS_W}}}t')
        if t_elem is None:
            t_elem = etree.SubElement(primer_run, f'{{{NS_W}}}t')
        t_elem.text = nuevo_texto
        t_elem.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
        for r in runs[1:]:
            for t in r.findall(f'{{{NS_W}}}t'):
                t.text = ''
    if img_bytes and img_ext:
        mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg',
                    '.jpeg': 'image/jpeg', '.bmp': 'image/bmp', '.tiff': 'image/tiff'}
        img_mime = mime_map.get(img_ext, 'image/png')
        img_internal = f'word/media/firma_img{img_ext}'
        img_target   = f'media/firma_img{img_ext}'
        modificados[img_internal] = img_bytes
        ct_root = etree.fromstring(archivos['[Content_Types].xml'])
        extension = img_ext.lstrip('.')
        existentes = {e.get('Extension') for e in ct_root.findall(f'{{{CT_NS}}}Default')}
        if extension not in existentes:
            nd = etree.SubElement(ct_root, f'{{{CT_NS}}}Default')
            nd.set('Extension', extension)
            nd.set('ContentType', img_mime)
        modificados['[Content_Types].xml'] = etree.tostring(
            ct_root, xml_declaration=True, encoding='UTF-8', standalone=True)
        rels_name = 'word/_rels/document.xml.rels'
        rels_root = etree.fromstring(archivos[rels_name]) if rels_name in archivos else etree.fromstring(
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>')
        rel_id = 'rFirmaImg1'
        ids_existentes = {e.get('Id') for e in rels_root.findall(f'{{{REL_NS}}}Relationship')}
        while rel_id in ids_existentes:
            rel_id += 'x'
        new_rel = etree.SubElement(rels_root, f'{{{REL_NS}}}Relationship')
        new_rel.set('Id', rel_id)
        new_rel.set('Type', REL_TYPE)
        new_rel.set('Target', img_target)
        modificados[rels_name] = etree.tostring(
            rels_root, xml_declaration=True, encoding='UTF-8', standalone=True)
        for docpr in root.iter(f'{{{NS_WP}}}docPr'):
            if '{{firma}}' not in (docpr.get('name') or ''):
                continue
            docpr.set('name', 'firma_img')
            anchor = docpr.getparent()
            while anchor is not None and etree.QName(anchor.tag).localname != 'anchor':
                anchor = anchor.getparent()
            if anchor is None:
                continue
            ext_el = anchor.find(f'{{{NS_WP}}}extent')
            cx = ext_el.get('cx', '1809750') if ext_el is not None else '1809750'
            cy = ext_el.get('cy', '923925')  if ext_el is not None else '923925'
            blip_encontrado = False
            for blip in anchor.iter(f'{{{NS_A}}}blip'):
                blip.set(f'{{{NS_R}}}embed', rel_id)
                blip_encontrado = True
                break
            if not blip_encontrado:
                for gd in anchor.iter(f'{{{NS_A}}}graphicData'):
                    for child in list(gd):
                        gd.remove(child)
                    gd.set('uri', 'http://schemas.openxmlformats.org/drawingml/2006/picture')
                    pic_xml = (
                        f'<pic:pic xmlns:pic="{NS_PIC}" xmlns:a="{NS_A}" xmlns:r="{NS_R}">'
                        f'<pic:nvPicPr>'
                        f'<pic:cNvPr id="99" name="firma_img"/>'
                        f'<pic:cNvPicPr/>'
                        f'</pic:nvPicPr>'
                        f'<pic:blipFill>'
                        f'<a:blip r:embed="{rel_id}"/>'
                        f'<a:stretch><a:fillRect/></a:stretch>'
                        f'</pic:blipFill>'
                        f'<pic:spPr>'
                        f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
                        f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                        f'</pic:spPr>'
                        f'</pic:pic>'
                    )
                    gd.append(etree.fromstring(pic_xml))
                    break
    modificados['word/document.xml'] = etree.tostring(
        root, xml_declaration=True, encoding='UTF-8', standalone=True)
    zin_buf.seek(0)
    with zipfile.ZipFile(zin_buf, 'r') as zin:
        return _zip_copy_preservando_formato(zin, modificados)


# ─────────────────────────────────────────────
#  PROCESADOR PRINCIPAL
# ─────────────────────────────────────────────

def construir_nombre_archivo(plantilla_nombre, datos_fila, campos_nombre_archivo):
    ext = os.path.splitext(plantilla_nombre)[1].lower()
    nombre_base = os.path.splitext(plantilla_nombre)[0]
    partes = []
    for campo in campos_nombre_archivo:
        valor = formatear_valor(datos_fila.get(campo, ""))
        valor_limpio = re.sub(r'[<>:"/\\|?*]', '_', valor).strip()
        if valor_limpio:
            partes.append(valor_limpio)
    sufijo = "_".join(partes) if partes else "sin_nombre"
    return f"{nombre_base}_{sufijo}{ext}"


def procesar_plantilla(plantilla_nombre, plantilla_bytes, datos_fila, firmas_dict,
                       campos_nombre_archivo, campo_firma="firma", log_fn=None):
    ext = os.path.splitext(plantilla_nombre)[1].lower()
    valor_firma_col = formatear_valor(datos_fila.get(campo_firma, ""))
    img_bytes, img_ext = encontrar_firma(valor_firma_col, firmas_dict, log_fn)
    nombre_archivo = construir_nombre_archivo(plantilla_nombre, datos_fila, campos_nombre_archivo)
    if ext == ".xlsx":
        resultado = llenar_xlsx_zip(plantilla_bytes, datos_fila)
        resultado = insertar_firmas_en_drawing_xlsx(resultado, img_bytes, img_ext, datos_fila, log_fn)
    elif ext == ".docx":
        resultado = llenar_docx(plantilla_bytes, datos_fila, img_bytes, img_ext)
    else:
        return None, None
    return nombre_archivo, resultado


# ─────────────────────────────────────────────
#  ESTILOS CSS PERSONALIZADOS
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Generador de Documentos",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Fondo general */
.stApp {
    background: #0f1117;
}

/* Header principal */
.app-header {
    background: linear-gradient(135deg, #1a1f2e 0%, #16213e 50%, #0f3460 100%);
    border: 1px solid rgba(99, 179, 237, 0.15);
    border-radius: 16px;
    padding: 32px 40px;
    margin-bottom: 28px;
    position: relative;
    overflow: hidden;
}
.app-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, rgba(99,179,237,0.08) 0%, transparent 70%);
    pointer-events: none;
}
.app-header h1 {
    color: #e2e8f0;
    font-size: 26px;
    font-weight: 600;
    margin: 0 0 6px 0;
    letter-spacing: -0.5px;
}
.app-header p {
    color: #718096;
    font-size: 14px;
    margin: 0;
    font-weight: 300;
}
.app-header .badge {
    display: inline-block;
    background: rgba(99,179,237,0.12);
    border: 1px solid rgba(99,179,237,0.3);
    color: #63b3ed;
    font-size: 11px;
    font-weight: 500;
    padding: 3px 10px;
    border-radius: 20px;
    margin-bottom: 12px;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}

/* Tarjetas de sección */
.section-card {
    background: #1a1f2e;
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 12px;
    padding: 24px;
    margin-bottom: 16px;
    transition: border-color 0.2s;
}
.section-card:hover {
    border-color: rgba(99,179,237,0.2);
}
.section-label {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 16px;
}
.section-number {
    background: linear-gradient(135deg, #3182ce, #63b3ed);
    color: white;
    width: 26px;
    height: 26px;
    border-radius: 8px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 13px;
    font-weight: 600;
    flex-shrink: 0;
}
.section-title {
    color: #e2e8f0;
    font-size: 15px;
    font-weight: 500;
}

/* Widgets de Streamlit */
.stFileUploader > div {
    background: #111827 !important;
    border: 1px dashed rgba(99,179,237,0.25) !important;
    border-radius: 10px !important;
    transition: border-color 0.2s !important;
}
.stFileUploader > div:hover {
    border-color: rgba(99,179,237,0.5) !important;
}
.stSelectbox > div > div,
.stMultiSelect > div > div {
    background: #111827 !important;
    border-color: rgba(255,255,255,0.1) !important;
    color: #e2e8f0 !important;
}
label {
    color: #a0aec0 !important;
    font-size: 13px !important;
    font-weight: 400 !important;
}

/* Botón primario */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #2b6cb0, #3182ce) !important;
    border: none !important;
    border-radius: 10px !important;
    color: white !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 15px !important;
    font-weight: 600 !important;
    padding: 14px 28px !important;
    letter-spacing: 0.3px !important;
    transition: all 0.2s !important;
    box-shadow: 0 4px 20px rgba(49,130,206,0.35) !important;
}
.stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #3182ce, #4299e1) !important;
    box-shadow: 0 6px 28px rgba(49,130,206,0.5) !important;
    transform: translateY(-1px) !important;
}

/* Botón descarga */
.stDownloadButton > button {
    background: linear-gradient(135deg, #276749, #38a169) !important;
    border: none !important;
    border-radius: 10px !important;
    color: white !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    padding: 12px 24px !important;
    box-shadow: 0 4px 20px rgba(56,161,105,0.35) !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover {
    background: linear-gradient(135deg, #38a169, #48bb78) !important;
    transform: translateY(-1px) !important;
}

/* Área de log */
.stCode {
    background: #0d1117 !important;
    border: 1px solid rgba(255,255,255,0.06) !important;
    border-radius: 10px !important;
    font-family: 'DM Mono', monospace !important;
    font-size: 12px !important;
}

/* Info / warning / success */
.stInfo {
    background: rgba(49,130,206,0.1) !important;
    border-color: rgba(49,130,206,0.3) !important;
    border-radius: 10px !important;
}
.stSuccess {
    background: rgba(56,161,105,0.1) !important;
    border-color: rgba(56,161,105,0.3) !important;
    border-radius: 10px !important;
}
.stWarning {
    background: rgba(237,137,54,0.1) !important;
    border-color: rgba(237,137,54,0.3) !important;
    border-radius: 10px !important;
}

/* Divisor */
hr {
    border-color: rgba(255,255,255,0.06) !important;
    margin: 24px 0 !important;
}

/* Dataframe */
.stDataFrame {
    border-radius: 10px !important;
    overflow: hidden !important;
}

/* Caption */
.stCaption {
    color: #4a5568 !important;
}

/* Stat chips */
.stat-chip {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    background: rgba(99,179,237,0.08);
    border: 1px solid rgba(99,179,237,0.2);
    color: #63b3ed;
    font-size: 12px;
    font-weight: 500;
    padding: 4px 12px;
    border-radius: 20px;
    margin: 4px 4px 0 0;
}

/* Expander */
.streamlit-expanderHeader {
    background: #1a1f2e !important;
    border-radius: 8px !important;
    color: #a0aec0 !important;
    font-size: 13px !important;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────────

st.markdown("""
<div class="app-header">
    <div class="badge">📄 Herramienta interna</div>
    <h1>Generador de Documentos</h1>
    <p>Rellena plantillas Excel y Word automáticamente desde tu base de datos</p>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  SECCIÓN 1 y 2 — en columnas
# ─────────────────────────────────────────────

col1, col2 = st.columns(2, gap="medium")

hoja_sel              = None
campo_firma           = "firma"
campos_nombre_archivo = []
encabezados           = []
datos_filas           = []

with col1:
    st.markdown("""
    <div class="section-label">
        <div class="section-number">1</div>
        <span class="section-title">Base de datos</span>
    </div>
    """, unsafe_allow_html=True)

    excel_file = st.file_uploader(
        "Sube el Excel con los datos",
        type=["xlsx", "xls", "xlsm"],
        label_visibility="collapsed",
        help="Archivo Excel con las filas de productores o registros a procesar"
    )

    if excel_file:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(excel_file.read()), read_only=True, data_only=True)
            hojas = wb.sheetnames
            wb.close()
            excel_file.seek(0)

            hoja_sel = st.selectbox("Hoja", hojas)

            excel_file.seek(0)
            wb2 = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
            ws2 = wb2[hoja_sel]
            filas_raw = list(ws2.iter_rows(values_only=True))
            wb2.close()
            excel_file.seek(0)

            if filas_raw:
                encabezados = [str(c).strip() if c is not None else f"col_{i}"
                               for i, c in enumerate(filas_raw[0])]
                datos_filas = [
                    {encabezados[i]: fila[i] for i in range(len(encabezados))}
                    for fila in filas_raw[1:]
                    if any(v is not None for v in fila)
                ]

                st.markdown(f"""
                <div style="margin-top:8px;">
                    <span class="stat-chip">✔ {len(datos_filas)} filas</span>
                    <span class="stat-chip">📋 {len(encabezados)} columnas</span>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

                firma_default_idx = next(
                    (i for i, c in enumerate(encabezados) if "firma" in c.lower()), 0
                )
                campo_firma = st.selectbox(
                    "Columna de firma",
                    options=encabezados,
                    index=firma_default_idx
                )

                campos_nombre_archivo = st.multiselect(
                    "Campos para nombre del archivo",
                    options=encabezados,
                    default=[encabezados[0]],
                    help="El nombre del archivo generado usará estos campos"
                )

                if campos_nombre_archivo and datos_filas:
                    ejemplo = "_".join(
                        re.sub(r'[<>:"/\\|?*]', "_", formatear_valor(datos_filas[0].get(c, ""))).strip()
                        for c in campos_nombre_archivo
                    )
                    st.caption(f"Ejemplo: `plantilla_{ejemplo}.xlsx`")

        except Exception as e:
            st.error(f"No se pudo leer el Excel: {e}")

with col2:
    st.markdown("""
    <div class="section-label">
        <div class="section-number">2</div>
        <span class="section-title">Plantillas</span>
    </div>
    """, unsafe_allow_html=True)

    plantillas_files = st.file_uploader(
        "Plantillas",
        type=["xlsx", "docx"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        help="Sube una o más plantillas con marcadores {{campo}}"
    )

    if plantillas_files:
        st.markdown(f"""
        <div style="margin-top:8px;">
            <span class="stat-chip">📎 {len(plantillas_files)} plantilla(s) cargada(s)</span>
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  SECCIÓN 3 — Firmas
# ─────────────────────────────────────────────

st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
st.markdown("""
<div class="section-label">
    <div class="section-number">3</div>
    <span class="section-title">Firmas <span style="color:#4a5568; font-size:13px; font-weight:300">(opcional)</span></span>
</div>
""", unsafe_allow_html=True)

firmas_files = st.file_uploader(
    "Firmas",
    type=["png", "jpg", "jpeg", "bmp", "tiff"],
    accept_multiple_files=True,
    label_visibility="collapsed",
    help="El nombre del archivo debe coincidir con el valor en la columna de firma del Excel"
)

if firmas_files:
    st.markdown(f"""
    <div style="margin-top:8px; margin-bottom:4px;">
        <span class="stat-chip">🖊 {len(firmas_files)} firma(s) cargada(s)</span>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  SECCIÓN 4 — Filtro
# ─────────────────────────────────────────────

datos_filtrados = datos_filas

if datos_filas and encabezados:
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    st.markdown("""
    <div class="section-label">
        <div class="section-number">4</div>
        <span class="section-title">Filtro <span style="color:#4a5568; font-size:13px; font-weight:300">(opcional)</span></span>
    </div>
    """, unsafe_allow_html=True)

    col_f1, col_f2 = st.columns([1, 2])
    with col_f1:
        usar_filtro = st.checkbox("Filtrar filas antes de generar")

    if usar_filtro:
        with col_f1:
            columna_filtro = st.selectbox("Columna a filtrar", encabezados)
        if columna_filtro:
            valores_unicos = sorted(
                set(
                    formatear_valor(fila.get(columna_filtro))
                    for fila in datos_filas
                    if fila.get(columna_filtro) is not None and str(fila.get(columna_filtro)).strip() != ""
                )
            )
            with col_f2:
                valores_sel = st.multiselect(
                    f"Valores de '{columna_filtro}'",
                    options=valores_unicos,
                    default=valores_unicos,
                )
            if valores_sel:
                datos_filtrados = [
                    fila for fila in datos_filas
                    if formatear_valor(fila.get(columna_filtro)) in valores_sel
                ]
                st.info(f"🔎 {len(datos_filtrados)} de {len(datos_filas)} filas seleccionadas")
            else:
                datos_filtrados = []
                st.warning("Sin valores seleccionados — no se generará ningún documento.")

    with st.expander(f"👁  Ver filas a procesar ({len(datos_filtrados)})", expanded=False):
        if datos_filtrados:
            import pandas as pd
            st.dataframe(pd.DataFrame(datos_filtrados), use_container_width=True)
        else:
            st.write("Sin filas.")


# ─────────────────────────────────────────────
#  BOTÓN GENERAR
# ─────────────────────────────────────────────

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
st.divider()

generar = st.button("⚡  GENERAR DOCUMENTOS", type="primary", use_container_width=True)

if generar:
    if not excel_file:
        st.warning("⚠ Sube el Excel de datos.")
    elif not hoja_sel:
        st.warning("⚠ Selecciona la hoja.")
    elif not plantillas_files:
        st.warning("⚠ Sube al menos una plantilla.")
    elif not datos_filtrados:
        st.warning("⚠ No hay filas para procesar.")
    else:
        log_area = st.empty()
        logs = []

        def log(msg):
            logs.append(msg)
            log_area.code("\n".join(logs), language=None)

        firmas_dict = {}
        for f in (firmas_files or []):
            firmas_dict[f.name] = f.read()

        try:
            log(f"✔ {len(datos_filtrados)} filas a procesar | Campos: {', '.join(encabezados)}")
            log(f"✔ Firma desde columna: '{campo_firma}'")
            log(f"✔ Nombre de archivo desde: {campos_nombre_archivo}")
            log(f"✔ Firmas cargadas: {list(firmas_dict.keys()) or 'ninguna'}")

            zip_buffer = io.BytesIO()
            total = 0

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zout:
                for plantilla_file in plantillas_files:
                    plantilla_nombre = plantilla_file.name
                    plantilla_bytes  = plantilla_file.read()
                    log(f"\n📄 {plantilla_nombre}")

                    nombres_usados = {}

                    for i, fila in enumerate(datos_filtrados):
                        try:
                            nombre_arch, contenido = procesar_plantilla(
                                plantilla_nombre, plantilla_bytes, fila,
                                firmas_dict, campos_nombre_archivo, campo_firma, log
                            )
                            if nombre_arch and contenido:
                                valor_firma = formatear_valor(fila.get(campo_firma, ""))
                                _, img_ext_found = encontrar_firma(valor_firma, firmas_dict)
                                subcarpeta = "con_firma" if img_ext_found else "sin_firma"

                                nombre_sin_ext = os.path.splitext(nombre_arch)[0]
                                ext_arch = os.path.splitext(nombre_arch)[1]
                                clave = f"{subcarpeta}/{nombre_arch}"
                                if clave in nombres_usados:
                                    nombres_usados[clave] += 1
                                    nombre_arch = f"{nombre_sin_ext} ({nombres_usados[clave]}){ext_arch}"
                                else:
                                    nombres_usados[clave] = 0

                                ruta_zip = f"{os.path.splitext(plantilla_nombre)[0]}/{subcarpeta}/{nombre_arch}"
                                zout.writestr(ruta_zip, contenido)
                                log(f"   ✔ {ruta_zip}")
                                total += 1
                        except Exception as e:
                            import traceback
                            log(f"   ❌ Fila {i+2}: {e}")
                            log(traceback.format_exc())

            log(f"\n✅ {total} archivos generados")

            st.success(f"✅ {total} documentos generados correctamente")
            st.download_button(
                label="📥  Descargar todos los documentos (.zip)",
                data=zip_buffer.getvalue(),
                file_name="documentos_generados.zip",
                mime="application/zip",
                use_container_width=True
            )

        except Exception as e:
            import traceback
            st.error(f"Error: {e}")
            st.code(traceback.format_exc())
